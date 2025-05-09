
import os
import re
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def extrair_info_cabecalho(texto):
    nome_match = re.search(r"Nome:\s*(.+)", texto)
    cartao_match = re.search(r"Número do cartão:\s*([Xx\.0-9]+)", texto)
    vencimento_match = re.search(r"(\d{2}/\d{2}/\d{4})\s+Data de vencimento", texto)
    if not vencimento_match:
        vencimento_match = re.search(r"Data de vencimento:\s*(\d{2}/\d{2}/\d{4})", texto)
    nome = nome_match.group(1).strip() if nome_match else ""
    cartao = cartao_match.group(1).strip() if cartao_match else ""
    vencimento = vencimento_match.group(1).strip() if vencimento_match else ""
    return nome, cartao, vencimento

def normalizar_valores(valor):
    valor = valor.replace(' ', '')
    return valor.replace('.', '').replace(',', '.')

def extrair_transacoes_com_total(pdf_path):
    transacoes = []
    nome = cartao = vencimento = ""
    total_usd = total_brl = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()
            if not texto:
                continue
            if page.page_number == 1:
                nome, cartao, vencimento = extrair_info_cabecalho(texto)
            linhas = texto.split('\n')
            for linha in linhas:
                if "Total:" in linha:
                    total_match = re.search(r"Total:\s+(-?\s?[\d.,]+)\s+(-?\s?[\d.,]+)", linha)
                    if total_match:
                        total_usd = normalizar_valores(total_match.group(1))
                        total_brl = normalizar_valores(total_match.group(2))
                    break
                match = re.match(r'^(\d{2}/\d{2})\s+(.+?)\s+(-?\s?[\d.,]+)\s+(-?\s?[\d.,]+)$', linha)
                if match:
                    data, historico, usd, brl = match.groups()
                    transacoes.append({
                        "Data": data,
                        "Histórico": historico.strip(),
                        "US$": normalizar_valores(usd),
                        "R$": normalizar_valores(brl)
                    })
    return transacoes, nome, cartao, vencimento, total_usd, total_brl

def criar_visao_geral_layout_melhorado(wb, planilhas_dados):
    aba_geral = wb.create_sheet(title="Visão Geral")
    linha_atual = 1
    header_font = Font(bold=True, size=13, color="FFFFFF")
    titulo_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    destaque_font = Font(bold=True, size=11)
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_titulo = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_total = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for nome_aba, df, nome, cartao, vencimento, total_usd, total_brl in planilhas_dados:
        aba_geral.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=4)
        cel = aba_geral.cell(row=linha_atual, column=1, value=nome)
        cel.font = header_font
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.fill = fill_header
        cel.border = thin_border
        linha_atual += 1
        info_labels = [("Número do Cartão:", cartao), ("Data de Vencimento:", vencimento)]
        for label, val in info_labels:
            aba_geral.append([label, val])
            aba_geral.cell(row=linha_atual, column=1).font = titulo_font
            aba_geral.cell(row=linha_atual, column=1).fill = fill_titulo
            linha_atual += 1
        linha_atual += 1
        colunas = ["Data", "Histórico", "US$", "R$"]
        for col, val in enumerate(colunas, start=1):
            cel = aba_geral.cell(row=linha_atual, column=col, value=val)
            cel.font = destaque_font
            cel.fill = fill_titulo
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = thin_border
        linha_atual += 1
        for _, row in df.iterrows():
            aba_geral.append([row["Data"], row["Histórico"], row["US$"], row["R$"]])
            for col in range(1, 5):
                aba_geral.cell(row=linha_atual, column=col).font = normal_font
                aba_geral.cell(row=linha_atual, column=col).border = thin_border
            linha_atual += 1
        aba_geral.append(["Total Geral", "", float(total_usd), float(total_brl)])
        for col in range(1, 5):
            cel = aba_geral.cell(row=linha_atual, column=col)
            cel.font = destaque_font
            cel.fill = fill_total
            cel.border = thin_border
        linha_atual += 3
    for col in range(1, 5):
        letra = get_column_letter(col)
        aba_geral.column_dimensions[letra].width = [12, 50, 12, 14][col - 1]
    return wb

def processar_pdfs_com_visao_geral(pasta_pdf, arquivo_saida_excel):
    wb = Workbook()
    wb.remove(wb.active)
    planilhas_dados = []
    for nome_arquivo in os.listdir(pasta_pdf):
        if nome_arquivo.lower().endswith(".pdf"):
            caminho_pdf = os.path.join(pasta_pdf, nome_arquivo)
            dados, nome, cartao, vencimento, total_usd, total_brl = extrair_transacoes_com_total(caminho_pdf)
            if not dados:
                continue
            df = pd.DataFrame(dados)
            df["US$"] = pd.to_numeric(df["US$"], errors='coerce')
            df["R$"] = pd.to_numeric(df["R$"], errors='coerce')
            aba = wb.create_sheet(title=os.path.splitext(nome_arquivo)[0][:31])
            aba.append(["Nome:", nome])
            aba.append(["Número do Cartão:", cartao])
            aba.append(["Data de Vencimento:", vencimento])
            aba.append([])
            for r in dataframe_to_rows(df, index=False, header=True):
                aba.append(r)
            aba.append([])
            aba.append(["Total Geral", "", float(total_usd), float(total_brl)])
            planilhas_dados.append((nome_arquivo, df, nome, cartao, vencimento, total_usd, total_brl))
    criar_visao_geral_layout_melhorado(wb, planilhas_dados)
    wb.save(arquivo_saida_excel)
